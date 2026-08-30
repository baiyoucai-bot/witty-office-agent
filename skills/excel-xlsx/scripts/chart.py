"""给 .xlsx 加原生图表和条件格式（openpyxl，沙箱预装）。

    <沙箱 Python> chart.py --input 表.xlsx --output 表-图.xlsx --spec 图.json
    <沙箱 Python> chart.py --help-spec

**必须放在最后一步跑**：openpyxl 重写整本工作簿，会丢掉输入文件里已有的
图表和图片，所以输入里已有图表/图片时拒绝，除非 --force。公式保得住。
加完照旧跑 check_xlsx.py --original 核一遍。

退出码 0 写出成功；2 spec 不合法 / 没有 openpyxl / 输入已有图表且没 --force。
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path

SPEC_HELP = """spec JSON：

{
  "charts": [
    {
      "sheet": "汇总",
      "type": "bar",                 // bar 柱状 | barh 条形 | line 折线 | pie 饼
      "title": "各项目金额",
      "categories": "A2:A5",         // 类别轴（可写 汇总!A2:A5）
      "series": [
        {"name": "金额", "values": "D2:D5"}
      ],
      "anchor": "F2"                 // 图表左上角落在哪个格
    }
  ],
  "conditional_formats": [
    {"sheet": "汇总", "range": "D2:D10", "type": "color_scale",
     "colors": ["F8696B", "FFEB84", "63BE7B"]},
    {"sheet": "汇总", "range": "D2:D10", "type": "data_bar", "color": "638EC6"},
    {"sheet": "汇总", "range": "D2:D10", "type": "cell_is",
     "operator": "greaterThan", "formula": "1000", "fill": "FFC7CE"},
    {"sheet": "汇总", "range": "D2:D10", "type": "formula",
     "formula": "$D2>AVERAGE($D$2:$D$10)", "fill": "FFEB9C"}
  ]
}

charts 和 conditional_formats 至少给一个。透视表和 VBA 做不了（openpyxl 无此能力）。
"""


def _openpyxl():
    try:
        import openpyxl
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
        from openpyxl.formatting.rule import (
            CellIsRule,
            ColorScaleRule,
            DataBarRule,
            FormulaRule,
        )
        from openpyxl.styles import PatternFill
        from openpyxl.utils import range_boundaries
    except ImportError:
        raise RuntimeError(
            "没有 openpyxl：用沙箱解释器跑，或 uv pip install --python <解释器> openpyxl"
        ) from None
    return (
        openpyxl,
        {"bar": BarChart, "barh": BarChart, "line": LineChart, "pie": PieChart},
        Reference,
        Series,
        {"cell_is": CellIsRule, "color_scale": ColorScaleRule,
         "data_bar": DataBarRule, "formula": FormulaRule},
        PatternFill,
        range_boundaries,
    )


def has_existing_drawing(path: Path) -> bool:
    with zipfile.ZipFile(path) as zin:
        return any(
            name.startswith(("xl/charts/", "xl/media/", "xl/drawings/"))
            for name in zin.namelist()
        )


def _split_ref(ref: str, default_sheet: str) -> tuple[str, str]:
    if "!" in ref:
        sheet, coords = ref.split("!", 1)
        return sheet.strip("'"), coords
    return default_sheet, ref


def _reference(Reference, range_boundaries, workbook, ref: str, default_sheet: str):
    sheet_name, coords = _split_ref(ref, default_sheet)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"表 {sheet_name} 不存在")
    min_col, min_row, max_col, max_row = range_boundaries(coords)
    return Reference(
        workbook[sheet_name],
        min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row,
    )


def add_charts(workbook, charts_spec: list, helpers) -> int:
    _, chart_types, Reference, Series, _, _, range_boundaries = helpers
    count = 0
    for index, item in enumerate(charts_spec, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"charts[{index}] 不是对象")
        sheet_name = str(item.get("sheet") or "")
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"charts[{index}].sheet「{sheet_name}」不存在")
        kind = str(item.get("type") or "bar")
        if kind not in chart_types:
            raise ValueError(f"charts[{index}].type 只认 {' '.join(sorted(chart_types))}")
        chart = chart_types[kind]()
        if kind == "bar":
            chart.type = "col"
        elif kind == "barh":
            chart.type = "bar"
        if item.get("title"):
            chart.title = str(item["title"])
        series_spec = item.get("series")
        if not isinstance(series_spec, list) or not series_spec:
            raise ValueError(f"charts[{index}].series 必须是非空数组")
        categories = None
        if item.get("categories"):
            categories = _reference(
                Reference, range_boundaries, workbook, str(item["categories"]), sheet_name
            )
        for s_index, entry in enumerate(series_spec, start=1):
            if not isinstance(entry, dict) or not entry.get("values"):
                raise ValueError(f"charts[{index}].series[{s_index}] 缺 values")
            values = _reference(
                Reference, range_boundaries, workbook, str(entry["values"]), sheet_name
            )
            series = Series(values, title=str(entry.get("name") or f"系列{s_index}"))
            chart.series.append(series)
        if categories is not None:
            chart.set_categories(categories)
        anchor = str(item.get("anchor") or "F2")
        workbook[sheet_name].add_chart(chart, anchor)
        count += 1
    return count


def add_conditional_formats(workbook, formats_spec: list, helpers) -> int:
    _, _, _, _, rules, PatternFill, _ = helpers
    count = 0
    for index, item in enumerate(formats_spec, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"conditional_formats[{index}] 不是对象")
        sheet_name = str(item.get("sheet") or "")
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"conditional_formats[{index}].sheet「{sheet_name}」不存在")
        cell_range = str(item.get("range") or "")
        if not cell_range:
            raise ValueError(f"conditional_formats[{index}] 缺 range")
        kind = str(item.get("type") or "")
        fill = None
        if item.get("fill"):
            color = str(item["fill"])
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        if kind == "color_scale":
            colors = [str(color) for color in item.get("colors") or []]
            if len(colors) == 2:
                rule = rules["color_scale"](
                    start_type="min", start_color=colors[0],
                    end_type="max", end_color=colors[1],
                )
            elif len(colors) == 3:
                rule = rules["color_scale"](
                    start_type="min", start_color=colors[0],
                    mid_type="percentile", mid_value=50, mid_color=colors[1],
                    end_type="max", end_color=colors[2],
                )
            else:
                raise ValueError(f"conditional_formats[{index}].colors 要 2 或 3 个颜色")
        elif kind == "data_bar":
            rule = rules["data_bar"](
                start_type="min", end_type="max",
                color=str(item.get("color") or "638EC6"),
            )
        elif kind == "cell_is":
            operator = str(item.get("operator") or "")
            formula = item.get("formula")
            if not operator or formula is None:
                raise ValueError(f"conditional_formats[{index}] cell_is 缺 operator/formula")
            formulas = [str(f) for f in (formula if isinstance(formula, list) else [formula])]
            rule = rules["cell_is"](operator=operator, formula=formulas, fill=fill)
        elif kind == "formula":
            if not item.get("formula"):
                raise ValueError(f"conditional_formats[{index}] formula 缺 formula")
            rule = rules["formula"](formula=[str(item["formula"])], fill=fill)
        else:
            raise ValueError(
                f"conditional_formats[{index}].type 只认 color_scale data_bar cell_is formula"
            )
        workbook[sheet_name].conditional_formatting.add(cell_range, rule)
        count += 1
    return count


def run(src: Path, dest: Path, spec: dict, force: bool) -> tuple[int, int]:
    helpers = _openpyxl()
    openpyxl = helpers[0]
    charts_spec = spec.get("charts") or []
    formats_spec = spec.get("conditional_formats") or []
    if not charts_spec and not formats_spec:
        raise ValueError("spec 里 charts / conditional_formats 至少给一个")
    if has_existing_drawing(src) and not force:
        raise ValueError(
            "输入里已有图表/图片/绘图层：openpyxl 重写会把它们丢掉。"
            "确认可丢再加 --force，或改成在生成工作簿的最后一步加图表"
        )
    workbook = openpyxl.load_workbook(src)
    charts = add_charts(workbook, charts_spec, helpers) if charts_spec else 0
    formats = add_conditional_formats(workbook, formats_spec, helpers) if formats_spec else 0
    dest.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(dest)
    return charts, formats


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="给 xlsx 加图表/条件格式")
    parser.add_argument("--input")
    parser.add_argument("--output")
    parser.add_argument("--spec")
    parser.add_argument("--force", action="store_true", help="输入已有图表时仍覆盖")
    parser.add_argument("--help-spec", action="store_true")
    args = parser.parse_args(argv)
    if args.help_spec:
        print(SPEC_HELP)
        return 0
    if not args.input or not args.output or not args.spec:
        print("需要 --input --output --spec，或 --help-spec", file=sys.stderr)
        return 2
    src = Path(args.input)
    spec_path = Path(args.spec)
    for path, label in ((src, "输入"), (spec_path, "spec")):
        if not path.is_file():
            print(f"找不到{label}: {path}", file=sys.stderr)
            return 2
    try:
        spec = json.loads(spec_path.read_text(encoding="utf-8"))
        charts, formats = run(src, Path(args.output), spec, args.force)
    except (json.JSONDecodeError, ValueError, RuntimeError, OSError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    print(f"{args.output}  charts={charts}  conditional_formats={formats}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
